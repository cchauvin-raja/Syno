# -*- coding: utf-8 -*-
##################
#### 3.7.2  ######
##################

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tqdm import tqdm
import re
import time
chemin = '/Users/cchauvin/Documents/Synonymes/'
xmls_a_traiter = {'JPG':['ST2V7_JPG_FR_SKU_20210219_030408.xml','20210218_FichierSynonymesSetup.xlsx','fr_FR'],\
    'KAL':['ST2V7_KAL_ES_SKU_20210222_010856.xml','SynonymesSetupKalamazoo.xlsx','es_ES']}
channel_a_traiter = 'KAL'
#channel =['JPG','MONDOFFICE','KALAMAZOO','BERNARD-FR','BERNARD-BE']
dico_locale ={'JPG':'fr_FR', 'MONDOFFICE':'it_IT','KAL':'es_ES','BERNARD-FR':'fr_FR','BERNARD-BE':['fr_BE','nl_BE']}
locale = dico_locale.get(channel_a_traiter)

ns = '{http://www.intershop.com/xml/ns/enfinity/7.0/xcs/impex}'
attriblang = '{http://www.w3.org/XML/1998/namespace}lang'

def ajout_keyword(soup,sku, mots, combien_de_mots_clés):     #truc = ajout_keyword(soupe, sku, keywordsenrichis, len(keywords_from_start))
    if combien_de_mots_clés == 0: #   faut créér la balise mots clés
        #precision = {"name" : "SearchIndexKeywords", "xml:lang" : "fr-FR"}
        precision = {"name" : "SearchIndexKeywords", "xml:lang" : "es-ES"}
        tagKeywords = sku.find('custom-attributes')
        valueNouveau = soup.new_tag('custom-attribute', attrs = precision)
        tagKeywords.append(valueNouveau)
        for mot in mots:
            ntag = soup.new_tag('value')
            ntag.string = mot 
            valueNouveau.insert(1,ntag)
    else:
        #tagKeywords = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "fr-FR"}) ## sur celui la je reajoute un tag <value> xx </value>
        tagKeywords = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "es-ES"}) ## sur celui la je reajoute un tag <value> xx </value>
        for mot in mots:
            valueNouveau = soup.new_tag('value')
            valueNouveau.string = mot
            tagKeywords.append(valueNouveau)
    success = True
    #valueNouveau = sku['sku'][0:2]+' '+sku['sku'][0:2]
    # tagKeywords.append(valueNouveau)
    return success


# def ajout_tag_keyword(soup,sku):
#     tagKeywords = sku.find('custom-attributes')
#     #valueNouveau = soup.new_tag('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "fr-FR"})
#     valueNouveau = soup.new_tag('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "es-ES"})
#     tagKeywords.append(valueNouveau)
#     return(soup)

# def cherchesynonyme(lecodesku, textes, les_synos):
#     enplus=[]
#     list_cle_valeur = les_synos[0]
#     dico_spe = les_synos[2]
#     for syno in les_synos[3]:  # les synos[3] est la liste des cles 
#         if syno[0] != '|': 
#             m = re.search(rf"['\.\s\-]+{syno}['\.\s\-]+",textes)
#             if m is not None:
#                 if type(list_cle_valeur.get(syno)) == list:
#                     for valeur in list_cle_valeur.get(syno):
#                         enplus.append(valeur)
#                 else:
#                     enplus.append(list_synonymes.get(syno))
#     if dico_spe.get(lecodesku) is not None: 
#         valeur_spe = dico_spe.get(lecodesku)
#         valeurs = valeur_spe.split('|')
#         for valeur in valeurs:
#             enplus.append(valeur)
        
        # conversion unité  '|([\d,.]+)\scl
    # m = re.search(r"\b(\d+[,.\d]*)\s*cl\b",texte)
    # if m is not None:
    #     contenancecl = float(m.group(1)) if '.' in m.group(1) else float(m.group(1).replace(',','.'))
    #     enplus.append(str(contenancecl/1000)+ 'l')
    #     enplus.append(str(contenancecl/1000)+ 'litres')
    #     enplus.append(str(contenancecl/1000)+ ' l')
    #     enplus.append(str(contenancecl/1000)+ ' litres')
    # new = set(enplus)
    # old = set(keywords_from_start)
    # enrichis = new.difference(old)
    # return enplus#

# def renvoie_les_chaines_a_analyser(sku):
#     # name lang - long-description lang - short-description lang
#     # balise_name = sku.find('name',{"xml:lang" : "fr-FR"})  ## preciser la langue
#     # balise_short = sku.find('short-description' ,{"xml:lang" : "fr-FR"} )
#     # balise_long = sku.find('long-description' ,{"xml:lang" : "fr-FR"} )
#     balise_name = sku.find('name',{"xml:lang" : "es-ES"})  ## preciser la langue
#     balise_short = sku.find('short-description' ,{"xml:lang" : "es-ES"} )
#     balise_long = sku.find('long-description' ,{"xml:lang" : "es-ES"} )
#     vname = balise_name.string.lower() if balise_name is not None else ' '
#     vshort = balise_short.string.lower() if balise_short is not None else ' '
#     vlong = balise_long.string.lower() if balise_long is not None else ' '
#     texte = ' ' + vname + ' '+vshort+' '+vlong
#     texte2 = re.sub(r'&lt;.*?&gt;', ' ', texte)
#     texte2 = re.sub(r'<[^<]+?>', ' ',texte2)
#     #texte2 = re.sub(r'&lt;.*?&gt;', ' ', texte)
#     return texte2

# def renvoie_keywords(sku):
#     keywords=[]
#     # tagKeywords = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "fr-FR"} )
#     tagKeywords = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "es-ES"} )

#     if tagKeywords is not None:
#         for value in tagKeywords.stripped_strings:
#             keywords.append(value)
#     return keywords


# def ouvre_et_charge_syno(fichier):
#     print(time.strftime('debut chargement synonyes : %H %M %S'))
#     dico_synonymes ={}
#     dico_application_syno = {}
#     dico_specific = {}
#     nomdefichier = chemin + fichier
#     wb=load_workbook(nomdefichier)
#     sheet=wb.active
#     # qmax_lignes=sheet.max_row
#     i=2
#     cle = sheet.cell(row=i,column=1).value.lower()
#     while cle is not None:
#         cle = sheet.cell(row=i,column=1).value.lower()
#         cle = cle.replace('\/','/')
#         valeurs = sheet.cell(row=i,column=2).value.lower()
#         valeur = valeurs.split('|')
#         dico_synonymes[cle]=valeur
#         application = sheet.cell(row=i,column=3).value.lower()
#         dico_application_syno[cle]=application
#         i+=1
#         cle = sheet.cell(row=i,column=1).value
    
#     i = 2
#     cle_temp= sheet.cell(row=i,column=5).value
#     while cle_temp is not None:
#         cle_temp= sheet.cell(row=i,column=5).value
#         dico_specific[cle_temp] = str(sheet.cell(row=i,column=6).value).lower()
#         i+=1
#     print(time.strftime('fin chargement synonyes : %H %M %S'))
#     return [dico_synonymes,dico_application_syno,dico_specific]

# def renvoie_source_light(fichier):
#     print(time.strftime('debut nettoyage du fichier : %H %M %S'))
#     txt = chemin + fichier
#     xml = txt + 'good.xml'
#     with open(txt, 'r') as f_input:
#         lignes = f_input.readlines()
#         with open(xml, "w") as f_output:
#             for ligne in lignes:
#                 supp = 0
#                 if re.search(r'''<custom-attribute name.*\/>''',ligne): # l'attribut est vide
#                     supp=1         
#                 if re.search(r'''<custom-attribute.*<\/custom-attribute>''',ligne) is not None:
#                     if re.search(r'''"Staples\.S\d\d|"EcoTax|"Eclass\s|"ShortLink|"Image|"Pictos|"Stock''', ligne) is not None:
#                         supp = 1
#                 if supp == 0:
#                     f_output.write(str(ligne))
#     print(time.strftime('fin nettoyage du fichier : %H %M %S'))
#     return xml ### le fichier raccourci 

def main():
    print(time.strftime('debut des opérations : %H %M %S - ')+ channel_a_traiter)
    print(time.strftime('debut nettoyage du XML : %H %M %S'))  #######################################################################
    xmlOrigine = chemin + xmls_a_traiter.get(channel_a_traiter)[0]  # je récupère fichier à traiter et j'allege
    xmlAllege = xmlOrigine + 'good.xml'
    with open(xmlOrigine, 'r') as f_input:
        lignes = f_input.readlines()
        with open(xmlAllege, "w") as f_output:
            for ligne in lignes:
                supp = 0
                if re.search(r'''<custom-attribute name.*\/>''',ligne): # si l'attribut est vide
                    supp=1         
                if re.search(r'''<custom-attribute.*<\/custom-attribute>''',ligne) is not None:
                    if re.search(r'''"Staples\.S\d\d|"EcoTax|"Eclass\s|"ShortLink|"Image|"Pictos|"Stock''', ligne) is not None:     ### si la ligne définit un de ces attributs on rajoutera typology
                        supp = 1
                if supp == 0:
                    f_output.write(str(ligne))
    print(time.strftime('fin nettoyage du XML : %H %M %S'))    # fin allegement
    print(time.strftime('debut chargement synonyes : %H %M %S'))
    dico_synonymes, dico_application_syno, dico_specific = {},{},{}
    nomdefichier = chemin + xmls_a_traiter.get(channel_a_traiter)[1]
    wb=load_workbook(nomdefichier)
    sheet=wb.active
    i=2
    cle = sheet.cell(row=i,column=1).value.lower()
    while cle is not None:
        cle = sheet.cell(row=i,column=1).value.lower()
        cle = cle.replace('\/','/')
        valeurs = sheet.cell(row=i,column=2).value.lower()
        valeur = valeurs.split('|')
        dico_synonymes[cle]=valeur
        application = sheet.cell(row=i,column=3).value.lower()
        dico_application_syno[cle]=application
        i+=1
        cle = sheet.cell(row=i,column=1).value
    i = 2
    cle_temp= sheet.cell(row=i,column=5).value
    while cle_temp is not None:
        cle_temp= sheet.cell(row=i,column=5).value
        dico_specific[cle_temp] = str(sheet.cell(row=i,column=6).value).lower()
        i+=1
    print(time.strftime('fin chargement synonyes : %H %M %S'))
    #  j'ai 3 dicos : dico_synonymes,dico_application_syno,dico_specific
    with open (xmlAllege) as file_to_parse:
        print(time.strftime('debut de parsing : %H %M %S'))
        soup = BeautifulSoup(file_to_parse, 'html.parser')
        print(time.strftime('fin de parsing : %H %M %S'))
        gardetrace = open('logSynonymes.txt','a')
        for sku in tqdm(soup.find_all('product')):   # sku est le noeud le code est (sku['sku'])
            lecodesku = sku['sku']
            textes, keywordsenrichis = ' ', set()
            balise_name = sku.find('name',{"xml:lang" : "es-ES"})  ## preciser la langue
            balise_short = sku.find('short-description' ,{"xml:lang" : "es-ES"} )
            balise_long = sku.find('long-description' ,{"xml:lang" : "es-ES"} )
            vname = balise_name.string.lower() if balise_name is not None else ' '
            vshort = balise_short.string.lower() if balise_short is not None else ' '
            vlong = balise_long.string.lower() if balise_long is not None else ' '
            texte = ' ' + vname + ' '+vshort+' '+vlong
            texte2 = re.sub(r'&lt;.*?&gt;', ' ', texte)
            textes = re.sub(r'<[^<]+?>', ' ',texte2)
            ###
            newKeywords=[]
            for syno in list(dico_synonymes.keys()):  # définit les synonymes sur terme 
                m = re.search(rf"['\.\s\-]+{syno}['\.\s\-]+",textes)
                if m is not None:
                    if type(dico_synonymes.get(syno)) == list:
                        for valeur in dico_synonymes.get(syno):
                            newKeywords.append(valeur)
                    else:
                        newKeywords.append(dico_synonymes.get(syno))
            ###
            if dico_specific.get(lecodesku) is not None:   # définit les synonymes specifiques à rune ref
                valeur_spe = dico_specific.get(lecodesku)
                valeurs = valeur_spe.split('|')
                for valeur in valeurs:
                    newKeywords.append(valeur)
            ###
            if len(newKeywords)>0:
                keywordsExistants =[]
                tagKeywordsExistants = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "es-ES"} )
                if tagKeywordsExistants is not None:
                    for value in tagKeywordsExistants.stripped_strings:
                        keywordsExistants.append(value)
                new = set(newKeywords)
                old = set(keywordsExistants)
                keywordsEnrichis = new.difference(old)
                if len(keywordsEnrichis) > 0 :
                    if len(keywordsExistants) == 0: #   faut créér la balise mots clés
                        #precision = {"name" : "SearchIndexKeywords", "xml:lang" : "fr-FR"}
                        precision = {"name" : "SearchIndexKeywords", "xml:lang" : "es-ES"}
                        tagKeywords = sku.find('custom-attributes')
                        valueNouveau = soup.new_tag('custom-attribute', attrs = precision)
                        tagKeywords.append(valueNouveau)
                        for mot in keywordsEnrichis:
                            ntag = soup.new_tag('value')
                            ntag.string = mot 
                            valueNouveau.insert(1,ntag)
                    else:
                        tagKeywords = sku.find('custom-attribute', {"name" : "SearchIndexKeywords"},{"xml:lang" : "es-ES"}) ## sur celui la je reajoute un tag <value> xx </value>
                        for mot in keywordsenrichis:
                            valueNouveau = soup.new_tag('value')
                            valueNouveau.string = mot
                            tagKeywords.append(valueNouveau)

                gardetrace.write(sku['sku']+ ' : '+ '|'.join(keywordsEnrichis)+'\n')
            else:
                gardetrace.write(sku['sku']+ ' supprimée\n')
                sku.decompose()     ### on vire la sku de la soupe pas besoin de réimporter
        gardetrace.close()
        print(time.strftime('fin ajout des mots clés : %H %M %S'))
    outxml = xmls_a_traiter.get(channel_a_traiter)[0] +'enrichi.xml'
    with open(outxml, "w") as f_output:
        f_output.write(str(soup))
print(time.strftime('fin écriture : %H %M %S'))
 
if __name__ == "__main__":
    main()